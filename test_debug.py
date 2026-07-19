import openpyxl

dir_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化"
path_data = f"{dir_path}\\スタジアム_データ.xlsx"

wb = openpyxl.load_workbook(path_data)
# idx 3 is 【データ】蓄積用
ws = wb.worksheets[3]
print("=== Scanning sheet:", ws.title)

from openpyxl.worksheet.formula import ArrayFormula

formula_cells = []
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        val = ws.cell(r, c).value
        if isinstance(val, ArrayFormula):
            val = val.text
        if isinstance(val, str) and val.startswith("="):
            col_letter = openpyxl.utils.get_column_letter(c)
            formula_cells.append(f"{col_letter}{r}: {val}")

print(f"Found {len(formula_cells)} formula cells in this sheet:")
for fc in formula_cells[:30]:
    print("  ", fc)
if len(formula_cells) > 30:
    print(f"  ... and {len(formula_cells) - 30} more.")

wb.close()
