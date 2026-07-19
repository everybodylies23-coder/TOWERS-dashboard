import openpyxl
import excel_calc_logic
import datetime

excel_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化\スタジアム_データ.xlsx"

def fix_all():
    print(f"Opening {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['【データ】蓄積用']
    
    print("Building maps...")
    specs_dict, settings_dict = excel_calc_logic.build_maps(wb)
    
    # Find real max row
    real_max_row = 1
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, 1).value is not None:
            real_max_row = r
            break
    
    print(f"Total data rows: {real_max_row - 1}")
    
    # Build all row data from A-G columns
    all_rows = []
    for r in range(2, real_max_row + 1):
        d_val = ws.cell(r, 1).value
        m_num = ws.cell(r, 3).value
        m_name = ws.cell(r, 2).value  # Column B: machine name
        if d_val and m_num:
            d_val = excel_calc_logic._to_datetime(d_val)
            if d_val is not None:
                all_rows.append({
                    "row_num": r,
                    "date": d_val,
                    "machine_name": str(m_name) if m_name else "",
                    "machine_number": int(m_num),
                    "g_games": excel_calc_logic._safe_int(ws.cell(r, 4).value),
                    "diff_coins": excel_calc_logic._safe_int(ws.cell(r, 5).value),
                    "bb_count": excel_calc_logic._safe_int(ws.cell(r, 6).value),
                    "rb_count": excel_calc_logic._safe_int(ws.cell(r, 7).value)
                })
    
    # Sort chronologically
    all_rows.sort(key=lambda x: (x["date"], x["row_num"]))
    
    print(f"Recalculating K-AF for all {len(all_rows)} rows...")
    processed_history = []
    count = 0
    jug_count = 0
    for row_data in all_rows:
        curr_r = row_data["row_num"]
        # Recalculate probabilities (Cols 8, 9, 10)
        games_val = row_data["g_games"]
        bb_val = row_data["bb_count"]
        rb_val = row_data["rb_count"]
        ws.cell(curr_r, 8, excel_calc_logic.format_probability(games_val, bb_val + rb_val))
        ws.cell(curr_r, 9, excel_calc_logic.format_probability(games_val, bb_val))
        ws.cell(curr_r, 10, excel_calc_logic.format_probability(games_val, rb_val))
        
        excel_calc_logic.calculate_kaf_for_row(ws, curr_r, row_data, settings_dict, specs_dict, processed_history)
        if "ジャグラー" in row_data["machine_name"]:
            jug_count += 1
        count += 1
        if count % 500 == 0:
            print(f"  ...{count}/{len(all_rows)} rows processed")
    
    print(f"All {count} rows calculated ({jug_count} Juggler rows detected). Saving...")
    while True:
        try:
            wb.save(excel_path)
            break
        except PermissionError:
            print("Excelファイルが開かれています。閉じてからENTERを押してください。")
            input()
            
    print("完了！全行のK列〜AF列が正しく再計算されました。")

if __name__ == "__main__":
    fix_all()
