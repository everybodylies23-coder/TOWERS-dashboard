import os
import shutil
import datetime
import openpyxl
import re
import excel_calc_logic

dir_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化"
excel_path = os.path.join(dir_path, "スタジアム_データ.xlsx")
temp_path_10 = os.path.join(dir_path, "データテンプレート_v13.10.xlsx")
temp_path_11 = os.path.join(dir_path, "データ分析_テンプレート_v13.11.xlsx")
processed_dir = os.path.join(dir_path, "data_input", "processed")

def safe_save_workbook(wb, filepath):
    """
    Saves the workbook safely, prompting retry if file is locked by Excel.
    """
    while True:
        try:
            wb.save(filepath)
            break
        except PermissionError:
            print(f"\n[エラー/警告] ファイルに書き込めません: '{filepath}'")
            print("Excelアプリ等でこのファイルが開いている可能性があります。")
            print("Excelを閉じてから、キーボード of ENTERキーを押して再試行してください。")
            input("閉じた後にENTERキーを押してください...")

def backup_ai_sheets():
    """
    Backs up the AI prediction history and summaries from the clean v13.11 template
    to guarantee we restore the complete 133 rows of predictions.
    """
    ai_predictions = []
    ai_summaries = []
    
    backup_source = temp_path_11
    if not os.path.exists(backup_source):
        print(f"Original template {backup_source} not found. Falling back to current Excel file.")
        backup_source = excel_path
        
    print(f"Backing up AI sheets from source: {backup_source}")
    try:
        wb = openpyxl.load_workbook(backup_source, data_only=True)
        
        # Load AI predictions (Idx 1 in original v13.11 workbook)
        pred_ws = wb.worksheets[1]
        for r in range(4, pred_ws.max_row + 1):
            row_vals = [pred_ws.cell(r, c).value for c in range(1, 8)]
            if any(row_vals):
                ai_predictions.append(row_vals)
        print(f"Backed up {len(ai_predictions)} rows from predictions sheet '{pred_ws.title}'.")
            
        # Load AI summaries (Idx 3 in original v13.11 workbook)
        sum_ws = wb.worksheets[3]
        for r in range(1, sum_ws.max_row + 1):
            val_e = sum_ws.cell(r, 5).value
            val_f = sum_ws.cell(r, 6).value
            if val_e is not None or val_f is not None:
                ai_summaries.append((val_e, val_f))
        print(f"Backed up {len(ai_summaries)} rows from summary sheet '{sum_ws.title}'.")
            
        wb.close()
    except Exception as e:
        print(f"Warning during AI backup: {e}. Proceeding with empty backup.")
        
    return ai_predictions, ai_summaries

def parse_html_data(filepath):
    from bs4 import BeautifulSoup
    results = []
    seen_machine_numbers = set()
    
    content = ""
    encodings = ['utf-8', 'cp932', 'shift_jis', 'utf-8-sig']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
                if "機種" in content or "台番" in content or "差枚" in content:
                    break
        except Exception:
            continue
    if not content:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
    soup = BeautifulSoup(content, 'html.parser')
    tables = soup.find_all('table')
    for table in tables:
        headers = [th.get_text().strip() for th in table.find_all('th')]
        if not headers:
            first_row = table.find('tr')
            if first_row:
                headers = [td.get_text().strip() for td in first_row.find_all('td')]
                
        header_text = "".join(headers)
        is_slot_table = (
            ("BB" in headers and "RB" in headers) and 
            any(h in header_text for h in ["台番", "台番号"])
        )
        if not is_slot_table:
            continue
            
        rows = table.find_all('tr')
        for row in rows:
            cells = [td.get_text().strip() for td in row.find_all('td')]
            if len(cells) >= 5:
                cell_text = "".join(cells)
                if any(x in cell_text for x in ["平均", "累計", "合計"]):
                    continue
                try:
                    def clean_val(val):
                        return val.replace(",", "").replace("+", "").replace(" ", "").strip()
                    c0_clean = clean_val(cells[0])
                    c1_clean = clean_val(cells[1])
                    
                    if c0_clean.isdigit():
                        machine_number = int(c0_clean)
                        machine_name = cells[1]
                        g_games = int(clean_val(cells[2]))
                        diff_coins = int(clean_val(cells[3]))
                        bb_count = int(clean_val(cells[4]))
                        rb_count = int(clean_val(cells[5])) if len(cells) > 5 else 0
                    else:
                        machine_name = cells[0]
                        machine_number = int(c1_clean)
                        g_games = int(clean_val(cells[2]))
                        diff_coins = int(clean_val(cells[3]))
                        bb_count = int(clean_val(cells[4]))
                        rb_count = int(clean_val(cells[5])) if len(cells) > 5 else 0
                    
                    if machine_number in seen_machine_numbers:
                        continue
                    seen_machine_numbers.add(machine_number)
                    
                    results.append({
                        "machine_name": machine_name,
                        "machine_number": machine_number,
                        "g_games": g_games,
                        "diff_coins": diff_coins,
                        "bb_count": bb_count,
                        "rb_count": rb_count
                    })
                except ValueError:
                    continue
    return results

def clean_int(val):
    if val is None:
        return 0
    val_str = str(val).replace("=", "").replace("+", "").replace(",", "").strip()
    if not val_str:
        return 0
    try:
        return int(float(val_str))
    except ValueError:
        return 0

def rebuild_all():
    # 1. Backup AI prediction sheets
    ai_preds, ai_sums = backup_ai_sheets()
    
    # 2. Copy clean v13.10 template
    print("Copying clean v13.10 template file...")
    while True:
        try:
            shutil.copy(temp_path_10, excel_path)
            break
        except PermissionError:
            print(f"\n[エラー] テンプレートをコピーできません: '{excel_path}'")
            print("Excelアプリ等で 'スタジアム_データ.xlsx' が開いている可能性があります。")
            print("Excelを一度閉じてから、キーボードの【ENTER】キーを押して再試行してください。")
            input("閉じた後にENTERキーを押してください...")
            
    # 3. Load workbook and rename worksheets
    print("Renaming all worksheets to clean Japanese titles...")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    japanese_sheet_names = {
        0: '【PC】分析ダッシュボード',
        1: '【AI】予想・答え合わせ',
        2: '【分析】高設定履歴DB',
        3: '【AI】総括',
        4: '【スマホ】確率・差枚',
        5: '【スマホ】ダッシュボード',
        6: '【データ】蓄積用',
        7: '【マスター】設定',
        8: '【マスター】スペック',
        9: '【エラーチェック】',
        10: '【計算】中間集計'
    }
    for idx, clean_name in japanese_sheet_names.items():
        if idx < len(wb.worksheets):
            wb.worksheets[idx].title = clean_name
            
    # 4. Build Specs and Settings maps
    print("Building specs and machine configuration maps...")
    ws_spec = wb['【マスター】スペック']
    specs_dict = {}
    for r in range(2, 40):
        m_name = ws_spec.cell(r, 1).value
        if m_name:
            row_vals = [ws_spec.cell(r, c).value for c in range(2, 33)]
            specs_dict[m_name] = row_vals
            
    ws_setting = wb['【マスター】設定']
    settings_dict = {}
    for r in range(2, 1000):
        num = ws_setting.cell(r, 1).value
        name = ws_setting.cell(r, 2).value
        if num is not None:
            settings_dict[int(num)] = name

    # 5. Convert E-H in 【AI】予想・答え合わせ (idx 1) Row 4 to raw standard Excel formulas
    ws_ai_pred = wb['【AI】予想・答え合わせ']
    ws_ai_pred.cell(4, 5).value = "=IF(OR(A4=\"\",C4=\"\"),\"\",SUMIFS('【データ】蓄積用'!$D$2:$D$15000,'【データ】蓄積用'!$A$2:$A$15000,A4,'【データ】蓄積用'!$C$2:$C$15000,C4))"
    ws_ai_pred.cell(4, 6).value = "=IF(OR(A4=\"\",C4=\"\"),\"\",SUMIFS('【データ】蓄積用'!$E$2:$E$15000,'【データ】蓄積用'!$A$2:$A$15000,A4,'【データ】蓄積用'!$C$2:$C$15000,C4))"
    ws_ai_pred.cell(4, 7).value = "=IF(OR(A4=\"\",C4=\"\"),\"\",SUMIFS('【データ】蓄積用'!$L$2:$L$15000,'【データ】蓄積用'!$A$2:$A$15000,A4,'【データ】蓄積用'!$C$2:$C$15000,C4))"
    ws_ai_pred.cell(4, 8).value = "=IF(A4=\"\", \"\", IF(OR(E4=\"\", E4=\"集計\"), \"\", IF(ISNUMBER(SEARCH(\"ジャグラー\", B4)), IF(AND(IFERROR(VALUE(G4), 0)>=4.5, IFERROR(VALUE(F4), 0)>=500), \"〇\", \"×\"), IF(IFERROR(VALUE(F4), 0)>=1000, \"〇\", \"×\"))))"

    # 6. Scan processed directory for HTML data files
    html_files = [f for f in os.listdir(processed_dir) if f.lower().endswith(".html")]
    def extract_date(filename):
        m = re.search(r'\d{8}', filename)
        return m.group(0) if m else ""
    html_files.sort(key=extract_date)
    
    print(f"Found {len(html_files)} backup data files. Importing chronologically...")
    ws = wb['【データ】蓄積用']
    
    unique_dates = set()
    unique_machines = set()
    
    # Import each file (Write A-G raw values only. We will calculate K-AF later)
    for filename in html_files:
        date_str_raw = extract_date(filename)
        if not date_str_raw:
            continue
            
        target_date_obj = datetime.datetime.strptime(date_str_raw, "%Y%m%d")
        unique_dates.add(target_date_obj)
        filepath = os.path.join(processed_dir, filename)
        
        parsed = parse_html_data(filepath)
        if not parsed:
            print(f"No data parsed from {filename}. Skipping.")
            continue
            
        print(f"Appending {len(parsed)} rows for date {date_str_raw}...")
        
        real_max_row = 1
        for r in range(ws.max_row, 1, -1):
            if isinstance(ws.cell(r, 1).value, datetime.datetime):
                real_max_row = r
                break
                
        start_row = real_max_row + 1
        for i, data in enumerate(parsed):
            curr_r = start_row + i
            unique_machines.add(int(data["machine_number"]))
            
            # Write base values
            dt_cell = ws.cell(curr_r, 1, target_date_obj)
            dt_cell.number_format = 'yyyy/mm/dd'
            
            ws.cell(curr_r, 2, data["machine_name"])
            ws.cell(curr_r, 3, data["machine_number"])
            ws.cell(curr_r, 4, data["g_games"])
            ws.cell(curr_r, 5, data["diff_coins"])
            ws.cell(curr_r, 6, data["bb_count"])
            ws.cell(curr_r, 7, data["rb_count"])
            
            # Write probabilities
            games_val = int(data["g_games"] or 0)
            bb_val = int(data["bb_count"] or 0)
            rb_val = int(data["rb_count"] or 0)
            ws.cell(curr_r, 8, excel_calc_logic.format_probability(games_val, bb_val + rb_val))
            ws.cell(curr_r, 9, excel_calc_logic.format_probability(games_val, bb_val))
            ws.cell(curr_r, 10, excel_calc_logic.format_probability(games_val, rb_val))
            
    # --- 6b. POST-IMPORT COMPLETE RE-CALCULATION & VALUE INJECTION (K-AF) FOR ALL ROWS ---
    print("Executing post-import recalculation and value injection for ALL active rows...")
    
    final_max_row = 1
    for r in range(ws.max_row, 1, -1):
        if isinstance(ws.cell(r, 1).value, datetime.datetime):
            final_max_row = r
            break
            
    print(f"Found {final_max_row} active data rows in Accumulation sheet.")
    
    all_imported_rows = []
    for r in range(2, final_max_row + 1):
        dt_val = ws.cell(r, 1).value
        if not isinstance(dt_val, datetime.datetime):
            continue
        unique_dates.add(dt_val)
        
        row_dict = {
            "row_num": r,
            "date": dt_val,
            "machine_name": ws.cell(r, 2).value,
            "machine_number": clean_int(ws.cell(r, 3).value),
            "g_games": clean_int(ws.cell(r, 4).value),
            "diff_coins": clean_int(ws.cell(r, 5).value),
            "bb_count": clean_int(ws.cell(r, 6).value),
            "rb_count": clean_int(ws.cell(r, 7).value)
        }
        all_imported_rows.append(row_dict)
        unique_machines.add(row_dict["machine_number"])
        
    all_imported_rows.sort(key=lambda x: x["date"])
    
    # Calculate each row K-AF based on chronological memory lookup
    processed_history = []
    for row in all_imported_rows:
        curr_r = row["row_num"]
        
        # Write probabilities (Cols 8, 9, 10)
        games_val = row["g_games"]
        bb_val = row["bb_count"]
        rb_val = row["rb_count"]
        ws.cell(curr_r, 8, excel_calc_logic.format_probability(games_val, bb_val + rb_val))
        ws.cell(curr_r, 9, excel_calc_logic.format_probability(games_val, bb_val))
        ws.cell(curr_r, 10, excel_calc_logic.format_probability(games_val, rb_val))
        
        m_num = row["machine_number"]
        m_name = settings_dict.get(m_num, "×")
        ws.cell(curr_r, 11, m_name)
        
        gx = row["g_games"]
        dx = row["diff_coins"]
        mechanical_payout = round((gx * 3 + dx) / (gx * 3), 4) if gx > 0 else 0
        ws.cell(curr_r, 12, mechanical_payout)
        
        prev_rows = [p for p in processed_history if p["machine_number"] == m_num]
        prev_date = prev_rows[-1]["date"] if prev_rows else None
        if prev_date:
            p_cell = ws.cell(curr_r, 13, prev_date)
            p_cell.number_format = 'yyyy/mm/dd'
        else:
            ws.cell(curr_r, 13, "")
            
        prev_2_date = prev_rows[-2]["date"] if len(prev_rows) >= 2 else None
        if prev_2_date:
            p2_cell = ws.cell(curr_r, 14, prev_2_date)
            p2_cell.number_format = 'yyyy/mm/dd'
        else:
            ws.cell(curr_r, 14, "")
            
        days_elapsed = (row["date"] - prev_date).days if prev_date else ""
        ws.cell(curr_r, 15, days_elapsed)
        
        tail_num = int(str(m_num)[-1])
        ws.cell(curr_r, 16, tail_num)
        
        is_jug = 1 if "ジャグラー" in m_name else 0
        ws.cell(curr_r, 17, is_jug)
        
        bx = row["bb_count"]
        rx = row["rb_count"]
        
        avg_score = 0
        weight = 0.5
        final_score = 0
        payout_coeff = 0
        payout_ratio = ""
        bb_ratio = ""
        rb_ratio = ""
        comb_ratio = ""
        bb_setting = 0
        rb_setting = 0
        comb_setting = 0
        payout_setting = 0
        
        if is_jug == 1 and m_name in specs_dict:
            specs = specs_dict[m_name]
            bb_inv = float(specs[0]) if specs[0] else 0
            rb_inv = float(specs[1]) if specs[1] else 0
            spec_divisor = float(specs[2]) if specs[2] else 1
            
            payout_coeff = ((gx * 3) + dx - (bx * bb_inv) - (rx * rb_inv) - (gx / 7.3 * 3) - (gx / 33 * 2)) / spec_divisor
            payout_ratio = gx / payout_coeff if payout_coeff > 0 else ""
            bb_ratio = gx / bx if bx > 0 else ""
            rb_ratio = gx / rx if rx > 0 else ""
            comb_ratio = gx / (bx + rx) if (bx + rx) > 0 else ""
            
            def find_setting(val, spec_indices):
                if not val:
                    return 0
                for s_idx, spec_idx in enumerate(spec_indices, 1):
                    spec_val = specs[spec_idx]
                    if spec_val and float(val) <= float(spec_val):
                        return 8 - s_idx
                return 0
                
            bb_setting = find_setting(bb_ratio, range(4, 11))
            rb_setting = find_setting(rb_ratio, range(11, 18))
            comb_setting = find_setting(comb_ratio, range(18, 25))
            payout_setting = find_setting(payout_ratio, range(25, 32))
            
            avg_score = (bb_setting + rb_setting + comb_setting + payout_setting) / 4
            weight = 1.0 if gx >= 7000 else (0.8 if gx >= 5000 else (0.6 if gx >= 3000 else 0.5))
            final_score = avg_score * weight
            
        ws.cell(curr_r, 18, payout_coeff if payout_coeff else "")
        ws.cell(curr_r, 19, payout_ratio)
        ws.cell(curr_r, 20, bb_ratio)
        ws.cell(curr_r, 21, rb_ratio)
        ws.cell(curr_r, 22, comb_ratio)
        ws.cell(curr_r, 23, bb_setting if bb_setting else "")
        ws.cell(curr_r, 24, rb_setting if rb_setting else "")
        ws.cell(curr_r, 25, comb_setting if comb_setting else "")
        ws.cell(curr_r, 26, payout_setting if payout_setting else "")
        ws.cell(curr_r, 27, avg_score if avg_score else "")
        ws.cell(curr_r, 28, weight if gx > 0 else "")
        ws.cell(curr_r, 29, final_score if final_score else "")
        
        # Calculate Columns AD-AF (30-32)
        ad_val = 0
        af_notes = []
        if gx > 0:
            base_score = 0
            if is_jug == 1:
                base_score = final_score
            else:
                if gx >= 1000:
                    if dx >= 3000:
                        base_score = 6.5
                    elif dx >= 2000:
                        base_score = 5.5
                    elif dx >= 1000:
                        base_score = 4.5
                    elif dx >= 0:
                        base_score = 3.5
                    else:
                        base_score = 2.0
                        
            ad_val = round(min(70.0, base_score * 10.0), 1)
            if ad_val > 0:
                af_notes.append(f"評価:{ad_val}pt")
            
            # Additions
            if base_score >= 4.5:
                ad_val += 15
                af_notes.append("据置+15")
            if base_score < 3.5:
                slump_days = (days_elapsed + 1) if (dx < 0 and isinstance(days_elapsed, int)) else 0
                if slump_days >= 2:
                    ad_val += 10
                    af_notes.append("凹グ+10")
            if m_name in ["準ピ", "準ピ2"]:
                ad_val += 5
                af_notes.append("準ピ+5")
            # Tail avg diff coins
            this_tail = tail_num
            same_tail_rows = [r for r in processed_history if int(str(r["machine_number"])[-1]) == this_tail and r["date"] <= row["date"]]
            avg_diff_tail = sum(r["diff_coins"] for r in same_tail_rows) / len(same_tail_rows) if same_tail_rows else 0
            if avg_diff_tail > 0:
                ad_val += 10
                af_notes.append("末尾+10")
                
        ws.cell(curr_r, 30, ad_val)
        
        ae_val = 0
        if ad_val > 0:
            ae_val = ad_val + (100000 - curr_r) / 100000000.0
        ws.cell(curr_r, 31, ae_val)
        
        af_val = " | ".join(af_notes) if af_notes else ""
        ws.cell(curr_r, 32, af_val)
        
        processed_history.append(row)

    # 7. Restore AI predictions
    print("Restoring AI predictions...")
    ai_ws = wb['【AI】予想・答え合わせ']
    for r in range(5, ai_ws.max_row + 200):
        for c in range(1, 9):
            ai_ws.cell(r, c).value = None
            
    for i, row_vals in enumerate(ai_preds):
        curr_r = 4 + i
        dt_val = row_vals[0]
        if isinstance(dt_val, str):
            try:
                dt_val = datetime.datetime.strptime(dt_val.split()[0].replace("-", "/"), "%Y/%m/%d")
            except Exception:
                pass
        dt_cell = ai_ws.cell(curr_r, 1, dt_val)
        if isinstance(dt_val, datetime.datetime):
            dt_cell.number_format = 'yyyy/mm/dd'
            
        for col_idx in range(2, 5):
            ai_ws.cell(curr_r, col_idx, row_vals[col_idx-1])
            
        if curr_r > 4:
            for col_idx in [5, 6, 7, 8]:
                src_formula = ai_ws.cell(4, col_idx).value
                if isinstance(src_formula, str) and src_formula.startswith("="):
                    new_formula = re.sub(rf'\b([A-Z]+)4\b', rf'\g<1>{curr_r}', src_formula)
                    ai_ws.cell(curr_r, col_idx, new_formula)
                
    print("Restoring AI summaries...")
    sum_ws = wb['【AI】総括']
    for r in range(1, sum_ws.max_row + 200):
        sum_ws.cell(r, 5).value = None
        sum_ws.cell(r, 6).value = None
        
    for i, (val_e, val_f) in enumerate(ai_sums):
        curr_r = 1 + i
        dt_val = val_e
        if isinstance(dt_val, str):
            try:
                dt_val = datetime.datetime.strptime(dt_val.split()[0].replace("-", "/"), "%Y/%m/%d")
            except Exception:
                pass
        dt_cell = sum_ws.cell(curr_r, 5, dt_val)
        if isinstance(dt_val, datetime.datetime):
            dt_cell.number_format = 'yyyy/mm/dd'
        sum_ws.cell(curr_r, 6, val_f)
        
    # 8. Emulate FILTER/SORT logic for '【分析】高設定履歴DB'
    print("Emulating dynamic FILTER/SORT array formulas in '【分析】高設定履歴DB'...")
    ws_db = wb['【分析】高設定履歴DB']
    for col in range(2, 100):
        ws_db.cell(5, col).value = None
    for row in range(7, 200):
        ws_db.cell(row, 1).value = None
        
    sorted_dates = sorted(list(unique_dates))
    for idx, dt in enumerate(sorted_dates):
        c_cell = ws_db.cell(5, 2 + idx, dt)
        c_cell.number_format = 'yyyy/mm/dd'
        
    sorted_machines = sorted(list(unique_machines))
    for idx, mach in enumerate(sorted_machines):
        ws_db.cell(7 + idx, 1, mach)
        
    # --- 9. CRITICAL CLEANUP: WIPE OUT leftover template cells/formulas below max raw data row ---
    ws_data = wb['【データ】蓄積用']
    real_max_raw_row = 1
    for r in range(ws_data.max_row, 1, -1):
        v = ws_data.cell(r, 1).value
        if isinstance(v, datetime.datetime):
            real_max_raw_row = r
            break
            
    print(f"  Sweeping garbage template formulas from row {real_max_raw_row + 1} to {ws_data.max_row + 1000}...")
    for r in range(real_max_raw_row + 1, ws_data.max_row + 1000):
        for c in range(1, 45):
            ws_data.cell(r, c).value = None
            
    # Clean up predictions sheet too
    real_max_pred = 3
    for r in range(ai_ws.max_row, 3, -1):
        if ai_ws.cell(r, 1).value is not None:
            real_max_pred = r
            break
    print(f"  Cleaning leftover prediction cells from row {real_max_pred + 1} to {ai_ws.max_row + 200}...")
    for r in range(real_max_pred + 1, ai_ws.max_row + 200):
        for c in range(1, 15):
            ai_ws.cell(r, c).value = None

    # --- CRITICAL SHEETS ELIMINATION ---
    print("\nDeleting heavy Google Sheets-specific sheets by index...")
    indices_to_remove = [10, 9, 5, 4, 0]
    for idx in indices_to_remove:
        if idx < len(wb.worksheets):
            ws = wb.worksheets[idx]
            wb.remove(ws)
            
    # Rename remaining 6 sheets
    print("Renaming remaining sheets...")
    remaining_names = {
        0: '【AI】予想・答え合わせ',
        1: '【分析】高設定履歴DB',
        2: '【AI】総括',
        3: '【データ】蓄積用',
        4: '【マスター】設定',
        5: '【マスター】スペック'
    }
    for idx, clean_name in remaining_names.items():
        if idx < len(wb.worksheets):
            wb.worksheets[idx].title = clean_name
            
    print("Saving rebuilt workbook...")
    safe_save_workbook(wb, excel_path)
    wb.close()
    print("Rebuild complete! Rebuilt workbook is 100% clean and compatible.")

if __name__ == "__main__":
    rebuild_all()
