import openpyxl

excel_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化\スタジアム_データ.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheetname in wb.sheetnames:
    if "履歴" in sheetname or "DB" in sheetname:
        ws = wb[sheetname]
        print(f"=== Sheet: {sheetname} ===")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
            print(f"  Row {r}: {vals}")

wb.close()
